"""sample-file.xlsx から既存データを取り込む初期インポート（任意）。

- 税金算出シート → tax_year_param（基礎/青色/税率上書き等）＋ tax_year_input（事業所得等）
- 生活費算出シート → monthly_entry（銀行 × 費目 × 年月 × 金額）

使い方（プロジェクトルート / PowerShell）:
    .venv/Scripts/python.exe -m scripts.import_excel                # 既定ファイル sample-file.xlsx
    .venv/Scripts/python.exe -m scripts.import_excel path\to.xlsx   # ファイル指定

冪等性：税は年度単位で upsert。生活費は取り込む年度の既存 monthly_entry を消してから再投入する。
生活費シートの合計/式・曖昧な列（合計列・W「りそな」等）は取り込まない（明細列のみ）。
実データを DB に書き込むため、実行はユーザーの任意。金額は生活費=円 / 税=万円。
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.config import BASE_DIR
from app.db import SessionLocal
from app.models import Bank, Category, MonthlyEntry, TaxBracket, TaxYearInput, TaxYearParam
from app.seeds.tax_masters import STANDARD_BRACKETS

# 生活費算出シート：列（アルファベット）→ (銀行名, 費目名, 種別)。合計列・曖昧列は含めない。
LIVING_COLUMN_MAP: dict[str, tuple[str, str, str]] = {
    "E": ("現金", "給与", "income"),
    "F": ("現金", "その他収入", "income"),
    "H": ("現金", "税金貯金", "saving"),
    "K": ("楽天銀行", "保険・ラテ等", "expense"),
    "L": ("楽天銀行", "くもん", "expense"),
    "M": ("楽天銀行", "NHK", "expense"),
    "N": ("楽天銀行", "iDeco", "saving"),
    "P": ("三井住友", "食費", "expense"),
    "Q": ("三井住友", "NISA", "saving"),
    "R": ("三井住友", "その他", "expense"),
    "S": ("三井住友", "電気", "expense"),
    "T": ("三井住友", "ガス", "expense"),
    "U": ("三井住友", "水道", "expense"),
    "V": ("三井住友", "保険", "expense"),
    "X": ("りそな", "コスモス", "expense"),
    "Y": ("りそな", "固定資産税", "expense"),
    "Z": ("りそな", "住民税", "expense"),
    "AA": ("りそな", "その他", "expense"),
    "AB": ("りそな", "アコム", "expense"),
    "AC": ("りそな", "プロミス", "expense"),
    "AD": ("りそな", "Forcus", "expense"),
    "AE": ("りそな", "小遣い", "expense"),
    "AG": ("UFJ", "国民年金", "expense"),
    "AH": ("UFJ", "小規模企業共済", "saving"),
    "AI": ("UFJ", "生命保険", "expense"),
    "AJ": ("UFJ", "国民健康保険", "expense"),
}

# 税金算出シート：列インデックス（1始まり）
TAX_COLS = {
    "year": 1, "business": 2, "salary": 3, "expenses": 4, "basic": 5, "spouse": 6,
    "life": 7, "social": 8, "small_biz": 9, "other": 10, "blue": 11, "earthquake": 12,
    "donation": 13, "rate": 14, "rate_deduction": 15, "reconstruction": 16,
    "resident_rate": 18, "resident_deduction": 19,
}


def _get_or_create_bank(session: Session, name: str, order: int) -> Bank:
    b = session.scalar(select(Bank).where(Bank.name == name))
    if b is None:
        b = Bank(name=name, display_order=order)
        session.add(b)
        session.flush()
    return b


def _get_or_create_category(session: Session, name: str, kind: str) -> Category:
    c = session.scalar(select(Category).where(Category.name == name))
    if c is None:
        c = Category(name=name, kind=kind, parent_id=None, display_order=100)
        session.add(c)
        session.flush()
    return c


def _num(value) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def import_tax(session: Session, wb) -> dict[str, int]:
    ws = wb["税金算出"]
    params_n = inputs_n = 0
    for row in range(1, ws.max_row + 1):
        year = ws.cell(row=row, column=TAX_COLS["year"]).value
        if not isinstance(year, int) or year < 2000:
            continue

        def cell(key):
            return _num(ws.cell(row=row, column=TAX_COLS[key]).value)

        param = session.get(TaxYearParam, year)
        if param is None:
            param = TaxYearParam(year=year)
            session.add(param)
            params_n += 1
        if cell("basic") is not None:
            param.basic_deduction_manyen = cell("basic")
        if cell("blue") is not None:
            param.blue_return_deduction_manyen = cell("blue")
        rate = cell("rate")
        if rate is not None:
            param.income_tax_rate_mode = "manual"
            param.income_tax_rate_override = rate
            param.income_tax_deduction_override = cell("rate_deduction")
        else:
            param.income_tax_rate_mode = "auto"
        if cell("reconstruction") is not None:
            param.reconstruction_tax_rate = cell("reconstruction")
        if cell("resident_rate") is not None:
            param.resident_tax_rate = cell("resident_rate")
        if cell("resident_deduction") is not None:
            param.resident_tax_deduction_manyen = cell("resident_deduction")

        # 累進表が無ければ標準表を投入
        if session.scalar(select(TaxBracket.id).where(TaxBracket.year == year).limit(1)) is None:
            for lo, up, r, ded in STANDARD_BRACKETS:
                session.add(TaxBracket(year=year, lower_bound_manyen=lo, upper_bound_manyen=up, rate=r, deduction_manyen=ded))

        ti = session.get(TaxYearInput, year)
        if ti is None:
            ti = TaxYearInput(year=year)
            session.add(ti)
            inputs_n += 1
        ti.business_income_manyen = cell("business") or 0
        ti.salary_income_manyen = cell("salary") or 0
        ti.expenses_manyen = cell("expenses") or 0
        ti.spouse_special_deduction_manyen = cell("spouse") or 0
        ti.life_insurance_deduction_manyen = cell("life") or 0
        ti.social_insurance_deduction_manyen = cell("social") or 0
        ti.small_biz_mutual_aid_deduction_manyen = cell("small_biz") or 0
        ti.other_income_deduction_manyen = cell("other") or 0
        ti.earthquake_insurance_deduction_manyen = cell("earthquake") or 0
        ti.donation_manyen = cell("donation") or 0

    session.commit()
    return {"params": params_n, "inputs": inputs_n}


def import_living(session: Session, wb) -> dict[str, int]:
    ws = wb["生活費算出"]
    # マスタ解決
    bank_cache: dict[str, Bank] = {}
    cat_cache: dict[str, Category] = {}
    order = 0
    for col, (bname, cname, kind) in LIVING_COLUMN_MAP.items():
        if bname not in bank_cache:
            bank_cache[bname] = _get_or_create_bank(session, bname, order)
            order += 1
        if cname not in cat_cache:
            cat_cache[cname] = _get_or_create_category(session, cname, kind)

    # 取り込む年度を把握し、既存 monthly_entry をクリア（冪等）
    years: set[int] = set()
    current_year = None
    rows_to_process: list[tuple[int, int, int]] = []  # (year, month, excel_row)
    for row in range(4, ws.max_row + 1):
        a = ws.cell(row=row, column=1).value
        b = ws.cell(row=row, column=2).value
        if isinstance(a, int):
            current_year = a
        if current_year is None or not isinstance(b, str) or not b.endswith("月"):
            continue
        try:
            month = int(b.replace("月", "").strip())
        except ValueError:
            continue
        if 1 <= month <= 12:
            years.add(current_year)
            rows_to_process.append((current_year, month, row))

    for y in years:
        session.execute(delete(MonthlyEntry).where(MonthlyEntry.year == y))

    entries_n = 0
    for year, month, row in rows_to_process:
        for col, (bname, cname, _kind) in LIVING_COLUMN_MAP.items():
            v = _num(ws.cell(row=row, column=column_index_from_string(col)).value)
            if v is None:
                continue
            amount = int(round(v / 10.0)) * 10  # 10円単位
            if amount <= 0:
                continue
            session.add(
                MonthlyEntry(
                    bank_id=bank_cache[bname].id,
                    category_id=cat_cache[cname].id,
                    year=year,
                    month=month,
                    amount_yen=amount,
                )
            )
            entries_n += 1

    session.commit()
    return {"years": sorted(years), "entries": entries_n}


def main() -> None:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(BASE_DIR) / "sample-file.xlsx"
    if not path.exists():
        print(f"ファイルが見つかりません: {path}")
        sys.exit(1)
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"取り込み元: {path}")
    with SessionLocal() as session:
        t = import_tax(session, wb)
        living = import_living(session, wb)
    print(f"税: tax_year_param +{t['params']} 件 / tax_year_input +{t['inputs']} 件（既存は更新）")
    print(f"生活費: 年度 {living['years']} / monthly_entry {living['entries']} 件")
    print("インポートが完了しました。")


if __name__ == "__main__":
    main()
